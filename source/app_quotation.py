import pandas as pd
import openpyxl as opxl
import math
import os as os
import sys
import json
import logging
from datetime import datetime
import math


# Function calculate total cost for all PORT and SHIPPING LINE
# args: No desc
# return: 2x (0-PORT, 1-SHIPPING, 2-TOTALCOST)
def get_quotation_data(file, sheet, skip) :
    # Load the Excel file into a pandas dataframe
    df = pd.read_excel(file, sheet_name=sheet, skiprows=skip)
    # Get data
    data20ft = df.iloc[:, [0,1,2,4,6,7,8,9,10,11]]
    data40ft = df.iloc[:, [0,1,3,5,6,7,8,9,10,12]]
    return data20ft, data40ft


# Function calculate total cost for all PORT and SHIPPING LINE
# data: 0-PORT, 1-SHIPPING, [2:9]-COST[...]
# return: 0-PORT, 1-SHIPPING, 2-TOTALCOST
def total_cost(data) :
    port = data.iloc[:,0].str.upper()
    shipping = data.iloc[:,1].str.upper()
    cost = data.iloc[:,2:10]
    # Convert cost to numberic data
    for i in range(len(cost.columns)):
        cost.isetitem(i, pd.to_numeric(cost.iloc[:,i], errors='coerce'))
    # drop row if main price is not provided or set to 0, not need to calculate
    cost.dropna(subset=[cost.columns[0]], inplace=True)
    cost = cost[cost[cost.columns[0]] != 0]
    # Calculate total cost
    total_cost = cost.sum(axis=1, numeric_only=True)
    # Return the list
    return pd.DataFrame().assign(PORT = port, SHIPPING = shipping, TOTALCOST = total_cost)


# Function return lowest cost provided by SHIPPING LINE for each PORT
# data: 0-PORT, 1-SHIPPING, 2-TOTALCOST
# return: 0-PORT, 1-SHIPPING, 2-TOTALCOST (only one min data for one port)
def min_shipping_line_cost(data) :
    #drop all NaN TOTALCOST value, not select these value, may lead to wrong result
    data.dropna(subset=[data.columns[2]], inplace=True)
    return data.loc[data.groupby(data.columns[0])[data.columns[2]].idxmin()]


# Function return data to fill for forwarder
# min_cost: 0-PORT, 1-SHIPPING, 2-TOTALCOST (only one min data for one port)
# return: idx, data(0-POD, 1-COST, 2-SHIPPING)
def prepare_forwarder_data(fwd_name, min_cost, file, sheet, skip) :
    # Load the Excel file into a pandas dataframe
    df = pd.read_excel(file, sheet_name=sheet, skiprows=skip)
    # Find fowarder index
    fwd_index = df.columns.get_loc(fwd_name)
    # Get data
    pod = df.loc[:, 'POD'].str.upper()
    dft = df.loc[:, fwd_name]
    # Make return frame
    fwd_data = pd.DataFrame().assign(POD = pod, COST = dft, SHIPPING = dft)
    # Create a dictionary to map PORT keys to TOTALCOST and SHIPPING values
    shipping = dict(zip(min_cost[min_cost.columns[0]], min_cost[min_cost.columns[1]]))
    cost = dict(zip(min_cost[min_cost.columns[0]], min_cost[min_cost.columns[2]]))
    # Use map function to fill columns
    fwd_data[fwd_data.columns[1]] = fwd_data[fwd_data.columns[0]].map(cost)
    fwd_data[fwd_data.columns[2]] = fwd_data[fwd_data.columns[0]].map(shipping)
    return fwd_index, fwd_data

# Function update data of forwarder to sheet
# fwd_data: 0-POD, 1-SHIPPING, 2-COST
# return: No desc
def write_forwarder_data(fwd_idx, fwd_data, file, sheet, skip) :
    # Load the existing Excel file
    workbook = opxl.load_workbook(file)
    worksheet = workbook[sheet]

    # Write the combined data to the worksheet, indexed #1
    for row in range(1, len(fwd_data)+1):
        cost = fwd_data[fwd_data.columns[1]].iloc[row-1]
        ship = fwd_data[fwd_data.columns[2]].iloc[row-1]
        if not math.isnan(cost):
            worksheet.cell(row=row+skip+1, column=fwd_idx+1).value = cost
            worksheet.cell(row=row+skip+1, column=fwd_idx+2).value = ship
            
    # Save the changes to the Excel file
    workbook.save(file)

    
# Function get configuration from file
# args: No desc
# return: No desc
def get_config(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            config = json.load(file)
    except Exception as e:
        return None, e
    
    return config, None

# Function printout progress in console
# progress: percentage
# return: No desc
def print_progress_bar(progress):
    bar_length = 100
    filled_length = int(bar_length * progress)
    bar = '[' + '#' * filled_length + ' ' * (bar_length - filled_length) + ']'
    sys.stdout.write('\r' + bar + ' %d%%' % (progress * 100))
    sys.stdout.flush()

# Function create template configuration file
# args: No desc
# return: No desc
def create_config_file(config_file_path):
    config_template = {
        "quotation":
        {
            "input_path" : "Path\\to\\folder\\contain\\your\\intput\\quotation\\files",
            "input_sheet" : "Input sheet name, all input files must be set same name",
            
            "area1" :
            {
                "process" : "Set value true/false (no quote \" \") to process",
                "suffix" : "suffix to detect quotation data file"
            },

            "area2" :
            {
                "process" : "Set value true/false (no quote \" \") to process",
                "suffix" : "suffix to detect quotation data file"
            }
        },

        "report" :
        {
            "output_path" : "Path\\to\\folder\\contain\\your\\report\\files",

            "area1" :
            {
                "report_file" : "Report file name",
                "20feet" :
                {
                    "report_sheet" : "Report sheet name",
                    "bestprices_sheet" : "Report sheet name for best price selection"
                },
                "40feet" :
                {
                    "report_sheet" : "Report sheet name",
                    "bestprices_sheet" : "Report sheet name for best price selection"
                }
            },

            "area2" :
            {
                "report_file" : "Report file name",
                "20feet" :
                {
                    "report_sheet" : "Report sheet name",
                    "bestprices_sheet" : "Report sheet name for best price selection"
                },
                "40feet" :
                {
                    "report_sheet" : "Report sheet name",
                    "bestprices_sheet" : "Report sheet name for best price selection"
                }
            }
        }
    }
    with open(config_file_path, "w") as config_file:
        json.dump(config_template, config_file, indent=4)

# Function return processed quotation data
# args: No desc
# return: No desc
def prepare_data_for_bestprices(file, sheet, skip_rows, skip_cols) :
    return pd.read_excel(file, sheet_name=sheet, skiprows=skip_rows).iloc[:, skip_cols:]
    

# Function return top best prices, default is top 4
# args: No desc
# return: No desc
def get_bestprices(data, top = 4):
    result = {}
    pods = pd.Series(data['POD'])
    for idx in range(len(pods)):
        vals = []
        # Data layout: 1-POD, [1,2] FWD1, [3,4] FWD2, ... 
        for i in range(1, len(data.columns), 2):
            fwd = data.columns[i]
            cost = data[data.columns[i]][idx]
            ship = data[data.columns[i+1]][idx]
            if not math.isnan(cost):
                vals.append((ship, fwd, cost))

        sorted_vals = sorted(vals, key=lambda x: x[2])
        sz = min(top, len(sorted_vals))
        result[pods[idx].upper()] = sorted_vals[:sz]
    return result

        
# Function fill top best prices in report
# args: No desc
# return: No desc
def write_bestprices_report(wdict, file, sheet, skip) :
    df = pd.read_excel(file, sheet_name=sheet, skiprows=skip)
    fwd_idx = df.columns.get_loc('DESTINATION')
    # Load the existing Excel file
    workbook = opxl.load_workbook(file)
    worksheet = workbook[sheet]
    # Write the combined data to the worksheet, indexed #1
    for row in range(1, len(df['DESTINATION'])+1):
        dest = worksheet.cell(row=row+skip+1, column=fwd_idx+1).value
        if dest in wdict and len(wdict[dest]) != 0:
            worksheet.cell(row=row+skip+1, column=fwd_idx+2).value = wdict[dest][0][0]
            worksheet.cell(row=row+skip+1, column=fwd_idx+3).value = wdict[dest][0][1]
            worksheet.cell(row=row+skip+1, column=fwd_idx+4).value = wdict[dest][0][2]
            del wdict[dest][0]
    # Save the changes to the Excel file
    workbook.save(file)

##########################################################################################
def process_quotation(config, logging):
    # [quotation]
    input_path = config['quotation']['input_path']
    input_sheet = config['quotation']['input_sheet']
    area1 = config['quotation']['area1']
    area2 = config['quotation']['area2']
    area1_suffix = config['quotation']['area1']['suffix']
    area2_suffix = config['quotation']['area2']['suffix']

    # [report]
    output_path = config['report']['output_path']
    area1_report_file = os.path.join(output_path, config['report']['area1']['report_file'])
    area2_report_file = os.path.join(output_path, config['report']['area2']['report_file'])
    area1_20feet_report_sheet = config['report']['area1']['20feet']['report_sheet']
    area1_40feet_report_sheet = config['report']['area1']['40feet']['report_sheet']

    if not os.path.exists(input_path):
        logging.error(f"{datetime.now()}: Input file path '{input_path}' not found")
        return False
    
    if area1 and not os.path.exists(area1_report_file):
        logging.error(f"{datetime.now()}: Report file '{area1_report_file}' not found")
        return False
    
    if area2 and not os.path.exists(area2_report_file):
        logging.error(f"{datetime.now()}: Report file '{area2_report_file}' not found")
        return False
    
    input_files = os.listdir(input_path)
    if not input_files:
        logging.info(f"{datetime.now()}: Input file path '{input_path}' is empty")
        return True

    # process
    try:
        processed = 0; total = len(input_files)
        for input in input_files:
            # progress bar update
            print_progress_bar(processed/total)
            processed += 1

            if area1 and (input.lower().endswith(area1_suffix + '.xls') or input.endswith(area1_suffix + '.xlsx')):
                output = area1_report_file
            elif area2 and (input.lower().endswith(area2_suffix + '.xls') or input.endswith(area2_suffix + '.xlsx')):
                output = area2_report_file
            else:
                continue    

            # Partner name
            name = input[:(input.find('.')-1)]
            
            # Get quotation information
            data20ft, data40ft = get_quotation_data(os.path.join(input_path, input), input_sheet, 0)

            # Calculate total cost for 20FT and 40FT
            sum20 = total_cost(data20ft)
            sum40 = total_cost(data40ft)

            # Sorting to get min cost information
            min_sum20 = min_shipping_line_cost(sum20)
            min_sum40 = min_shipping_line_cost(sum40)

            # Prepare data to fill for forwarder
            fwd_idx_20, fwd_data_20 = prepare_forwarder_data(name, min_sum20, output, area1_20feet_report_sheet, 3)
            fwd_idx_40, fwd_data_40 = prepare_forwarder_data(name, min_sum40, output, area1_40feet_report_sheet, 3)

            # Update data
            write_forwarder_data(fwd_idx_20, fwd_data_20, output, area1_20feet_report_sheet, 3)
            write_forwarder_data(fwd_idx_40, fwd_data_40, output, area1_40feet_report_sheet, 3)
        
        # final progress
        print_progress_bar(processed/total)
    except Exception as e:
        print_progress_bar(processed/total)
        logging.error(f"{datetime.now()}: Process quotation file {input} error: {str(e)}")
        return False
    
    return True


def best_prices(config, logging):
    output_path = config['report']['output_path']
    processed = 0
    try:
        for area in {"area1", "area2"} :
            report_file = os.path.join(output_path, config['report'][area]['report_file'])
            for size in {"20feet", "40feet"} :
                print_progress_bar(processed/4)
                processed += 1

                report_sheet = config['report'][area][size]['report_sheet']
                bestprices_sheet = config['report'][area][size]['bestprices_sheet']
                # process
                input = prepare_data_for_bestprices(report_file, report_sheet, 3, 2)
                wdict = get_bestprices(input, 4)
                write_bestprices_report(wdict, report_file, bestprices_sheet, 4)
        # final progress
        print_progress_bar(processed/4)
    except Exception as e:
        logging.error(f"{datetime.now()}: Build best price list {input} error: {str(e)}")
        return False
    
    return True

if __name__ == "__main__":
    # logging
    logging.basicConfig(filename='quotation.log', level=logging.INFO)
    # get configs
    config_file = 'config.json'
    if not os.path.exists(config_file):
        create_config_file(config_file)
        logging.error(f"{datetime.now()}: File configuration '{config_file}' not found, template file has been generated!")
        sys.exit()

    config, err = get_config(config_file)
    if err:
        logging.error(f"{datetime.now()}: File configuration '{config_file}' error: '{str(err)}'")
        sys.exit() 
    
    if process_quotation(config, logging):
        print("\r\nQuotation has been processed successfully, continue to select best price list!")
    else:
        print("\r\nQuotation has been processed fail, please check error in log file!")

    if best_prices(config, logging):
        print("\r\nBest price list has been processed successfully, please check report!")
    else:
        print("\r\nBest price list has been processed fail, please check error in log file!")

