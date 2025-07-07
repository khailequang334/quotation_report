import math
import pandas as pd
import openpyxl as opxl
import openpyxl.utils
from typing import Dict, Any, List


class ReportGenerator:
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
    
    def write_all_partners_data(self, partners_data: Dict[str, List], output_file: str) -> None:
        for container_type in ['20ft', '40ft']:
            if partners_data[container_type]:
                sheet_name = partners_data[container_type][0]['sheet']
                self._write_partners_to_sheet(partners_data[container_type], output_file, sheet_name, 3)
    
    def _copy_cell_formatting(self, source_cell, target_cell) -> None:
        if source_cell.font:
            target_cell.font = opxl.styles.Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )
        
        if source_cell.fill:
            target_cell.fill = opxl.styles.PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        if source_cell.border:
            target_cell.border = opxl.styles.Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        if source_cell.alignment:
            target_cell.alignment = opxl.styles.Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )

    def _set_column_widths(self, worksheet, start_col: int, partner_count: int) -> None:
        worksheet.column_dimensions['A'].width = 25
        
        pod_col_letter = opxl.utils.get_column_letter(start_col - 1)
        worksheet.column_dimensions[pod_col_letter].width = 22
        
        for i in range(2, start_col - 1):
            col_letter = opxl.utils.get_column_letter(i)
            worksheet.column_dimensions[col_letter].width = 15
        
        for i in range(partner_count):
            col_idx = start_col + i
            col_letter = opxl.utils.get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 18

    def _write_partners_to_sheet(self, partners_list: List[Dict], file: str, sheet: str, skip: int) -> None:
        workbook = opxl.load_workbook(file)
        worksheet = workbook[sheet]
        
        df = pd.read_excel(file, sheet_name=sheet, skiprows=skip)
        
        try:
            pod_col_idx = df.columns.get_loc('POD')
        except KeyError:
            pod_col_idx = 0
        
        start_col = pod_col_idx + 1
        header_row_pos = skip + 1
        pod_header_cell = worksheet.cell(row=header_row_pos, column=pod_col_idx + 1)
        
        for partner_idx, partner_info in enumerate(partners_list):
            partner_name = partner_info['partner']
            partner_data = partner_info['data']
            
            col_pos = start_col + partner_idx + 1
            
            partner_cell_1 = worksheet.cell(row=header_row_pos, column=col_pos)
            partner_cell_1.value = partner_name
            
            self._copy_cell_formatting(pod_header_cell, partner_cell_1)
            
            cost_mapping = dict(zip(partner_data['POD'], partner_data['COST']))
            
            for row in range(len(df)):
                pod = df.iloc[row]['POD']
                if pod in cost_mapping:
                    cost = cost_mapping[pod]
                    if not pd.isna(cost):
                        worksheet.cell(row=row + header_row_pos + 1, column=col_pos).value = cost
        
        self._set_column_widths(worksheet, start_col + 1, len(partners_list))
        workbook.save(file)
    
    def write_forwarder_data_to_file(self, forwarder_data: Dict[str, Any], output_file: str) -> None:
        for container_type in ['20ft', '40ft']:
            data_info = forwarder_data[container_type]
            partners_list = [{
                'partner': data_info['partner'],
                'data': data_info['data']
            }]
            self._write_partners_to_sheet(partners_list, output_file, data_info['sheet'], 3)

    def _write_single_forwarder_data(self, fwd_idx: int, fwd_data: pd.DataFrame, 
                                   file: str, sheet: str, skip: int) -> None:
        workbook = opxl.load_workbook(file)
        worksheet = workbook[sheet]

        for row in range(1, len(fwd_data) + 1):
            cost = round(float(fwd_data[fwd_data.columns[1]].iloc[row-1]), 2)
            if not math.isnan(cost):
                worksheet.cell(row=row + skip + 1, column=fwd_idx + 1).value = cost
                
        workbook.save(file)

    def generate_and_write_best_prices(self, report_file: str, area_config: Dict[str, Any]) -> None:
        for container_size in ['20feet', '40feet']:
            container_config = area_config[container_size]
            report_sheet = container_config['report_sheet']
            bestprices_sheet = container_config['bestprices_sheet']
            
            input_data = self._prepare_data_for_bestprices(report_file, report_sheet, 3, 2)
            best_prices_dict = self._get_best_prices(input_data, 4)
            self._write_bestprices_report(best_prices_dict, report_file, bestprices_sheet, 4)

    def _prepare_data_for_bestprices(self, file: str, sheet: str, 
                                   skip_rows: int, skip_cols: int) -> pd.DataFrame:
        return pd.read_excel(file, sheet_name=sheet, skiprows=skip_rows).iloc[:, skip_cols:]

    def _get_best_prices(self, data: pd.DataFrame, top: int = 4) -> Dict[str, list]:
        result = {}
        pods = pd.Series(data['POD'])
        
        for idx in range(len(pods)):
            vals = []
            for col_name in data.columns:
                if col_name != 'POD' and col_name not in ['DESTINATION', 'Unnamed: 0']:
                    try:
                        cost = data[col_name][idx]
                        if not pd.isna(cost) and not math.isnan(cost):
                            vals.append((col_name, cost))
                    except (ValueError, TypeError):
                        continue

            sorted_vals = sorted(vals, key=lambda x: x[1])
            sz = min(top, len(sorted_vals))
            result[pods[idx].upper()] = sorted_vals[:sz]
            
        return result

    def _write_bestprices_report(self, wdict: Dict[str, list], file: str, 
                               sheet: str, skip: int) -> None:
        df = pd.read_excel(file, sheet_name=sheet, skiprows=skip)
        fwd_idx = df.columns.get_loc('DESTINATION')
        
        workbook = opxl.load_workbook(file)
        worksheet = workbook[sheet]
        
        for row in range(1, len(df['DESTINATION'])+1):
            dest = worksheet.cell(row=row+skip+1, column=fwd_idx+1).value
            if dest in wdict and len(wdict[dest]) != 0:
                partner_name = wdict[dest][0][0]
                cost = round(float(wdict[dest][0][1]), 2)
                worksheet.cell(row=row+skip+1, column=fwd_idx+3).value = partner_name
                worksheet.cell(row=row+skip+1, column=fwd_idx+4).value = cost
                del wdict[dest][0]
        
        workbook.save(file) 